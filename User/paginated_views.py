from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from User.models import *

import openpyxl
from openpyxl.styles import Font, Alignment


ACCOUNTS_PER_PAGE = 8


def reset_paginated_Uticks(
    context: dict,
    omega_keywords: list[str]=["omegaUticked", "omega0Uticked"],
    total_keyword: str="allAccountsTicked",
    max_page_range: int=50,
) -> None:
  context[total_keyword] = []
  for i in range(1, max_page_range+1):
    for keyword in omega_keywords:
      context[keyword][i] = []


def paginate_account(
    request,
    global_context: dict,
    tick_keyword: str,          # Note: this is the keyword of obtained ticks
    untick_keyword: str,        # Note: this is the keyword of obtained unticks
    paginated_keyword: str,     # Note: this is the BACKEND of the paginated list
    total_keyword: str,         # Note: this is the keyword for all ticked users
    omega_keywords: list[str]=["omegaUticked", "omega0Uticked"],
    num_per_page: int=ACCOUNTS_PER_PAGE,
) -> dict:
  result = {}
  omega_tick_keyword, omega_untick_keyword = omega_keywords

  # Set up pagination with an amount per page
  all_accounts = Account.objects.all().order_by("dateSignUp")
  paginator = Paginator(all_accounts, num_per_page)

  # Get the current page number from the request
  page_num = request.GET.get('page', 1)
  page_obj = paginator.get_page(page_num)
  page_num_total = paginator.num_pages

  try:
    # Get the accounts for the current page
    accounts_page = paginator.page(page_num)
    current_pagenum = accounts_page.number
  except PageNotAnInteger:
    # If page is not an integer, deliver first page
    accounts_page = paginator.page(1)
    current_pagenum = 1
  except EmptyPage:
    # If page is out of range, deliver last page
    accounts_page = paginator.page(page_num_total)
    current_pagenum = page_num_total

  list_ticked = request.POST.get(tick_keyword, None)
  if not list_ticked:
    # If nothing found => no box has been ticked
    IDs_ticked = []
  else:
    # Capture the values and split into list (for current page)
    IDs_ticked = list(list_ticked.split(','))
    temp1, temp2 = [], []
    for userID in IDs_ticked:
      if userID not in temp1:
        temp2.append(userID)
        temp1.append(userID)
    # THE FIRST ELEMENT IS AN EMPTY (DONT KNOW WHY), THUS MUST BE REMOVED
    if '' in temp2:
      temp2.remove('')
    IDs_ticked = temp2

  list_unticked = request.POST.get(untick_keyword, None)
  if not list_unticked:
    # If nothing found => no box has been ticked
    IDs_unticked = []
  else:
    # Capture the values and split into list (for current page)
    IDs_unticked = list(list_unticked.split(','))
    temp1, temp2 = [], []
    for userID in IDs_unticked:
      if userID not in temp1:
        temp2.append(userID)
        temp1.append(userID)
    # THE FIRST ELEMENT IS AN EMPTY (DONT KNOW WHY), THUS MUST BE REMOVED
    if '' in temp2:
      temp2.remove('')
    IDs_unticked = temp2

  # Check the last click (tick or untick)
  last_click_status = request.POST.get("status")
  if last_click_status=="ticked":
    global_context[omega_tick_keyword][current_pagenum] = IDs_ticked
    global_context[omega_untick_keyword][current_pagenum] = [
      a.userID for a in accounts_page if a.userID not in IDs_ticked
    ]
  elif last_click_status=="unticked":
    global_context[omega_untick_keyword][current_pagenum] = IDs_unticked
    global_context[omega_tick_keyword][current_pagenum] = [
      a.userID for a in accounts_page if a.userID not in IDs_unticked
    ]

  result[total_keyword] = []
  for page in range(1, page_num_total+1):
    result[total_keyword].extend(global_context[omega_tick_keyword][page])
  result[paginated_keyword] = accounts_page
  return result


def paginate_account_Excel(
    request,
    context: dict,
    selections_keyword: str="allTickedAccounts",
) -> HttpResponse:
  response = HttpResponse(content_type="application/ms-excel")
  response["Content-Disposition"] = 'attachment; filename="account_list.xlsx"'

  wb = openpyxl.Workbook()
  sheet = wb.active
  titles = [
    "User ID",
    "Password",
    "Opr?",
    "Spl?",
    "Cpn?",
    "Joined In",
    "Company",
    "Supplier",
    "Email",
    "Vessel List"]
  column_widths = {
    'A': 15,
    'B': 15,
    'C': 10,
    'D': 10,
    'E': 10,
    'F': 15,
    'G': 20,
    'H': 20,
    'I': 20,
    'J': 40,
  }

  title_font = Font(bold=True, size=14)
  alignment = Alignment(horizontal='center', vertical='center')
  for col_letter, width in column_widths.items():
    sheet.column_dimensions[col_letter].width = width
  for i,title in enumerate(titles):
    cell = sheet.cell(row=1, column=i+1)
    cell.value = title
    cell.font = title_font
    cell.alignment = alignment

  accounts_ticked = context[selections_keyword]
  for j,account in enumerate(accounts_ticked):
    cell1 = sheet.cell(row=j+2, column=1)
    cell2 = sheet.cell(row=j+2, column=2)
    cell3 = sheet.cell(row=j+2, column=3)
    cell4 = sheet.cell(row=j+2, column=4)
    cell5 = sheet.cell(row=j+2, column=5)
    cell6 = sheet.cell(row=j+2, column=6)
    cell7 = sheet.cell(row=j+2, column=7)
    cell8 = sheet.cell(row=j+2, column=8)
    cell9 = sheet.cell(row=j+2, column=9)
    cell10 = sheet.cell(row=j+2, column=10)

    cell1.value = account.userID
    cell2.value = account.rawPassword
    cell3.value = account.isOpr
    cell4.value = account.isSpl
    cell5.value = account.isCpn
    cell6.value = account.dateSignUp
    cell7.value = account.companyName
    cell8.value = account.supplierName
    cell9.value = account.email
    cell10.value = account.vesselList

    for c in (cell1, cell2, cell3, cell4, cell5, cell6,
              cell7, cell8, cell9, cell10):
      c.alignment = alignment
  wb.save(response)
  return response


def paginate_account_delete(
    request,
    context: dict,
    selections_keyword: str="allTickedAccounts",
):
  accounts_ticked = context[selections_keyword]
  for account in accounts_ticked:
    account.delete()
