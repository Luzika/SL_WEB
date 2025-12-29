from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db.models import Q
from django.db.models.functions import Trim
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from Shipment.models import *
from Shipment.forms import *
from django.db import transaction
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import date #

import os
from django.conf import settings
from django.http import HttpResponse
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from datetime import datetime
import re
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import Spacer, Paragraph
from reportlab.lib import colors                  # ← Predefined colors: white, black, grey, etc.
from reportlab.lib.colors import HexColor         # ← For HexColor('#ff0000')
from reportlab.lib.enums import TA_LEFT, TA_CENTER


import logging

logger = logging.getLogger(__name__)
SHIPMENTS_PER_PAGE = 250

def reset_paginated_Sticks(
    context: dict,
    omega_keywords: list[str]=["shipmentTicked", "shipment0Ticked"],
    total_keyword: str="allShipmentsTicked",
    max_page_range: int=800,
) -> None:
  context[total_keyword] = []
  for i in range(1, max_page_range+1):
    for keyword in omega_keywords:
      context[keyword][i] = []

def filter_shipment(
    request,
    form_keyword: str,
    paginated_keyword: str,
    result_keyword: str,
    num_per_page: int=SHIPMENTS_PER_PAGE,
    filter_reload: django_filters.FilterSet=ShipmentFilter,
) -> dict:
    # Set up base queryset based on user type
    if request.user.isCpn:
        shipments = Shipment.objects.filter(company=request.user.companyName)
    elif request.user.isSpl:  # Check if the user is a supplier
        shipments = Shipment.objects.filter(supplier=request.user.supplierName)
    else:
        # User is staff/superuser: get ALL shipments
        shipments = Shipment.objects.all()

    # Apply filters
    params = request.GET.copy()
    removed_flag_vals = None
    if 'flag_status' in params:
        removed_flag_vals = params.getlist('flag_status')
        try:
            params.pop('flag_status')
        except Exception:
            try:
                del params['flag_status']
            except Exception:
                pass

    filter_form = filter_reload(params, queryset=shipments)
    filtered_queryset = filter_form.qs  # Get the queryset AFTER filtering (without flag_status)

    # Manually apply flag_status using the new function
    if removed_flag_vals:
        try:
            filtered_queryset = apply_flag_status_filter(filtered_queryset, removed_flag_vals)
        except Exception:
            # If helper missing or fails, ignore and continue
            pass

    # (debug logging removed)

    # --- Explicitly handle `flag_status` GET params to avoid django-filter string-iteration bug
    try:
        vals = request.GET.getlist('flag_status')
        if vals:
            # Annotate a trimmed version of the field to match empty or whitespace-only values
            try:
                from django.db.models.functions import Trim
                filtered_queryset = filtered_queryset.annotate(_trimmed_flag=Trim('flag_status'))
            except Exception:
                # If Trim is unavailable for any reason, proceed without it
                pass

            # Build Q to include BLANK semantics (empty, whitespace-only or NULL) when user selects 'BLANK'
            q = Q()
            non_blank = [v for v in vals if v != 'BLANK']
            if 'BLANK' in vals:
                q |= Q(flag_status__exact='BLANK')
                q |= Q(_trimmed_flag='')
                q |= Q(flag_status__isnull=True)
            if non_blank:
                # Match against the trimmed value to avoid mismatches due to stray spaces
                q |= Q(_trimmed_flag__in=non_blank)
            filtered_queryset = filtered_queryset.filter(q)
    except Exception:
        pass

    # === SERVER-SIDE SORTING ===
    sort_key = request.GET.get('sort')
    order = request.GET.get('order', 'asc')

    allowed_sort_fields = [
        'number', 'shipment_id', 'company', 'vessel', 'supplier', 'division', 
        'quanty', 'c_packing', 'weight', 'unit', 'size', 'docs', 'in_date', 
        'warehouse', 'by', 'IMP_BL', 'port', 'out_date', 'remark', 'EXP_BL', 
        'flag_status', 'insert_org', 'correct_org', 'Order_No'
    ]

    if sort_key and sort_key in allowed_sort_fields:
        prefix = '-' if order == 'desc' else ''
        filtered_queryset = filtered_queryset.order_by(f'{prefix}{sort_key}')
    else:
        # Default: Show LATEST (most recent) In Date at the top
        # Then newest shipment number as tie-breaker
        filtered_queryset = filtered_queryset.order_by('-in_date', '-number')

    # (final debug logging removed)

    # --- PAGINATION LOGIC ---
    paginator = Paginator(filtered_queryset, num_per_page)
    page_num = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
        
    current_pagenum = page_obj.number
    page_num_total = paginator.num_pages

    result = {
        form_keyword: filter_form,
        paginated_keyword: page_obj,
        result_keyword: filtered_queryset,
        "page_num": current_pagenum,
        "page_num_total": page_num_total,
        # Add for template: current sort state
        'current_sort': sort_key,
        'current_order': order,
    }
    return result

def paginate_shipment_PDF(
    request,
    context: dict,
    selections_keyword: str = "allTickedShipments",
) -> HttpResponse:
    shipments_ticked = context.get(selections_keyword, [])
    if not shipments_ticked:
        return HttpResponse("No data to export", status=400)

    # === 1. REGISTER KOREAN FONT ===
    font_path = os.path.join(settings.BASE_DIR, 'static_files', 'fonts', 'NanumGothic.ttf')
    font_name = 'Helvetica'
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('NanumGothic', font_path))
            font_name = 'NanumGothic'
            logger.info(f"Font successfully registered: {font_name}")
        except Exception as e:
            logger.error(f"Error registering font: {e}")
    else:
        logger.error(f"Font file NOT found at: {font_path}")

    # === 2. COLOR MAPPING ===
    COLOR_NAME_TO_HEX = {
        'black': '#000000', 'red': '#ff0000', 'green': '#008000', 'blue': '#0000ff',
        'orange': '#ffa500', 'purple': '#800080', 'brown': '#a52a2a',
        'gray': '#808080', 'grey': '#808080', 'pink': '#ffc0cb',
        'yellow': '#ffff00', 'white': '#ffffff',
    }

    def normalize_color(color_value):
        if not color_value:
            return '#000000'
        color_value = str(color_value).strip().lower()
        if color_value.startswith('#') and len(color_value) in (4, 7):
            return color_value.upper()
        return COLOR_NAME_TO_HEX.get(color_value, '#000000')

    def to_paragraph(text, style, clean_size=False):
        if not text:
            return Paragraph('', style)

        # 1. Convert to string and escape XML special characters first
        # This prevents & from breaking the PDF, but keeps our later <br/> safe
        safe_text = str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        if clean_size:
            # Clean and format Size field
            lines = [line.strip() for line in safe_text.split('\n') if line.strip()]
            formatted_lines = []
            for line in lines:
                line = re.sub(r'\s*[xX]\s*', ' × ', line)
                line = re.sub(r'\s*-\s*', ' — ', line)
                formatted_lines.append(line.strip())
            # 2. Join with <br/> AFTER escaping
            safe_text = '<br/>'.join(formatted_lines)
        else:
            # 2. Replace newlines with <br/> AFTER escaping
            safe_text = safe_text.replace('\n', '<br/>')

        return Paragraph(safe_text, style)

    # === 4. SUMMARY DATA ===
    unique_vessels = set()
    total_packages = 0
    total_weight = 0.0

    for s in shipments_ticked:
        if s.vessel:
            unique_vessels.add(s.vessel.strip())
        if s.quanty:
            try:
                total_packages += int(float(s.quanty))
            except (ValueError, TypeError):
                pass
        if s.weight:
            try:
                cleaned_weight = str(s.weight).replace(',', '')
                total_weight += float(cleaned_weight)
            except (ValueError, TypeError):
                pass

    vessels_str = ', '.join(sorted(unique_vessels)) if unique_vessels else 'N/A'

    # === 5. DOCUMENT SETUP ===
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="shipment_list.pdf"'

    left_margin = right_margin = 40
    top_margin = bottom_margin = 20
    page_width, page_height = landscape(A4)

    doc = SimpleDocTemplate(
        response,
        pagesize=(page_width, page_height),
        rightMargin=right_margin,
        leftMargin=left_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin
    )
    elements = []

    # Base style for cells
    cell_style_base = ParagraphStyle(
        'CellStyleBase',
        fontName=font_name,
        fontSize=8,
        leading=11,           # increased for better line spacing
        alignment=1,          # Center
        wordWrap='CJK',
    )

    # === 6. BUILD TABLE DATA ===
    headers = [
        "Vessel", "DOC", "Order No", "Supplier", "QTY", "Unit", "Size",
        "Weight", "In Date", "W/H", "IMP BL", "Remark", "U.Cre"
    ]
    data = [headers]

    for s in shipments_ticked:
        row_color_hex = normalize_color(getattr(s, 'colordiv', None))
        row_style = ParagraphStyle(
            'RowStyle',
            parent=cell_style_base,
            textColor=colors.HexColor(row_color_hex),
        )

        data.append([
            to_paragraph(s.vessel, row_style),
            to_paragraph(s.docs, row_style),
            to_paragraph(s.Order_No, row_style),
            to_paragraph(s.supplier, row_style),
            to_paragraph(s.quanty or '0', row_style),
            to_paragraph(getattr(s, 'unit', ''), row_style),
            to_paragraph(s.size, row_style, clean_size=True),           # ← Cleaned + multi-line
            to_paragraph(s.weight, row_style),
            to_paragraph(s.in_date.strftime('%Y-%m-%d') if s.in_date else '', row_style),
            to_paragraph(s.warehouse, row_style),
            to_paragraph(s.IMP_BL, row_style),
            to_paragraph(s.remark, row_style),
            to_paragraph(getattr(s, 'insert_org', ''), row_style),
        ])

    # === 7. COLUMN WIDTHS ===
    printable_width = page_width - left_margin - right_margin
    ratios = [0.10, 0.08, 0.15, 0.09, 0.05, 0.05, 0.14, 0.07, 0.08, 0.05, 0.09, 0.08, 0.05]
    col_widths = [max(30, int(printable_width * r)) for r in ratios]

    # === 8. CREATE TABLE WITH MULTI-PAGE SPLITTING ===
    # repeatRows=1 ensures the blue header appears on every new page
    # splitByRow=True allows a very long 'Size' or 'Remark' to split across pages
    table = Table(data, colWidths=col_widths, repeatRows=1, splitByRow=True)

    table.hAlign = 'CENTER'

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#cfe2ff")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        
        # IMPORTANT: Change VALIGN to TOP. 
        # If a row splits across pages, 'MIDDLE' can cause the text to disappear 
        # or look strange on the second page.
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), 
        
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        
        # Add extra padding to make long text easier to read
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    # === 9. ADD TO DOCUMENT WITH OVERFLOW PROTECTION ===
    summary_style = ParagraphStyle('SummaryStyle', fontName=font_name, fontSize=10, leading=14)

    elements.append(Paragraph("<b>SHIPMENT EXPORT SUMMARY</b>", summary_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>Vessels:</b> {vessels_str}", summary_style))
    elements.append(Paragraph(f"<b>Total Packages:</b> {total_packages:,} pkgs", summary_style))
    elements.append(Paragraph(f"<b>Total Weight:</b> {total_weight:,.2f} kg", summary_style))
    elements.append(Paragraph(f"<b>Export Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
    elements.append(Spacer(1, 20))
    
    # Crucial: This setting allows rows to break across pages if they are too long
    elements.append(table)

    # Use the build method
    doc.build(elements)
    return response
    
def paginate_shipment(
    request,
    filter_result: dict,
    global_context: dict,
    tick_keyword: str,          # Note: this is the keyword of obtained ticks
    untick_keyword: str,        # Note: this is the keyword of obtained unticks
    paginated_keyword: str,     # Note: this is the BACKEND of the paginated list
    total_keyword: str,         # Note: this is the keyword for all ticked shipments
    shipment_keywords: list[str]=["shipmentTicked", "shipment0Ticked"],
    num_per_page: int=SHIPMENTS_PER_PAGE,
) -> dict:
  result = {}
  shipment_tick_keyword, shipment_untick_keyword = shipment_keywords
  result.update(filter_result)

  # Set up pagination with an amount per page
  shipments = Shipment.objects.all().order_by("-number")
  paginator = Paginator(shipments, num_per_page)
  page_num_total = paginator.num_pages

  # Get the current page number from the request
  page_num = request.GET.get('page', 1)
  page_num_total = paginator.num_pages

  try:
    # Get the shipments for the current page
    shipments_page = paginator.page(page_num)
    current_pagenum = shipments_page.number
  except PageNotAnInteger:
    # If page is not an integer, deliver first page
    shipments_page = paginator.page(1)
    current_pagenum = 1
  except EmptyPage:
    # If page is out of range, deliver last page
    shipments_page = paginator.page(page_num_total)
    current_pagenum = page_num_total

  list_ticked = request.POST.get(tick_keyword, None)
  if not list_ticked:
    # If nothing found => no box has been ticked
    IDs_ticked = []
  else:
    # Capture the values and split into list (for current page)
    IDs_ticked = list(list_ticked.split(','))
    IDs_ticked = IDs_ticked[1:]

  list_unticked = request.POST.get(untick_keyword, None)
  if not list_unticked:
    # If nothing found => no box has been ticked
    IDs_unticked = []
  else:
    # Capture the values and split into list (for current page)
    IDs_unticked = list(list_unticked.split(','))
    IDs_unticked = IDs_unticked[1:]


#   Check the last click (tick or untick)
  last_click_status = request.POST.get("statusS")
  if last_click_status=="ticked":
    global_context[shipment_tick_keyword][current_pagenum] = [
      int(id_ticked) for id_ticked in IDs_ticked
    ]
    global_context[shipment_untick_keyword][current_pagenum] = [
      s.number for s in shipments_page if str(s.number) not in IDs_ticked
    ]
  elif last_click_status=="unticked":
    global_context[shipment_untick_keyword][current_pagenum] = [
      int(id_unticked) for id_unticked in IDs_unticked
    ]
    global_context[shipment_tick_keyword][current_pagenum] = [
      int(id_ticked) for id_ticked in IDs_ticked if id_ticked not in IDs_unticked
    ]

  result[total_keyword] = []
  for page in range(1, page_num_total+1):
    result[total_keyword].extend(global_context[shipment_tick_keyword][page])
#   if ("addShipment" in request.POST or "addShipmentM" in request.POST \
#     or "adjustShipment" in request.POST or "modifyShipment" in request.POST\
#     or "deleteShipment" in request.POST):
#     reset_paginated_Sticks(global_context, max_page_range=800)

  return result


def paginate_shipment_Excel(
    request,
    context: dict,
    selections_keyword: str = "allTickedShipments",
) -> HttpResponse:
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="shipment_list.xlsx"'

    wb = openpyxl.Workbook()
    sheet = wb.active

    shipments_ticked = context.get(selections_keyword, [])
    if not shipments_ticked:
        sheet.cell(row=1, column=1, value="No shipments selected")
        wb.save(response)
        return response

    # === CALCULATE SUMMARY ===
    unique_vessels = set()
    total_packages = 0
    total_weight = 0.0

    for shipment in shipments_ticked:
        if shipment.vessel:
            unique_vessels.add(shipment.vessel.strip())
        if shipment.quanty:
            try:
                total_packages += int(shipment.quanty)
            except (ValueError, TypeError):
                pass
        if shipment.weight:
            try:
                cleaned_weight = str(shipment.weight).replace(',', '')
                total_weight += float(cleaned_weight)
            except (ValueError, TypeError):
                pass

    vessels_str = ', '.join(sorted(unique_vessels)) if unique_vessels else 'N/A'

    # === SUMMARY SECTION ===
    summary_start_row = 1
    bold_font = Font(bold=True, size=14)
    normal_font = Font(size=12)

    # Title
    sheet.cell(row=summary_start_row, column=1, value="EXPORT SUMMARY")
    sheet.cell(row=summary_start_row, column=1).font = bold_font
    sheet.merge_cells(start_row=summary_start_row, start_column=1, end_row=summary_start_row, end_column=8)
    summary_start_row += 2

    # Vessels
    sheet.cell(row=summary_start_row, column=1, value=f"Vessels: {vessels_str}")
    sheet.cell(row=summary_start_row, column=1).font = normal_font
    sheet.merge_cells(start_row=summary_start_row, start_column=1, end_row=summary_start_row, end_column=8)
    summary_start_row += 1

    # Total Packages (new row)
    sheet.cell(row=summary_start_row, column=1, value=f"Total Packages: {total_packages:,} pkgs")
    sheet.cell(row=summary_start_row, column=1).font = normal_font
    sheet.merge_cells(start_row=summary_start_row, start_column=1, end_row=summary_start_row, end_column=8)
    summary_start_row += 1

    # Total Weight (next row)
    sheet.cell(row=summary_start_row, column=1, value=f"Total Weight: {total_weight:,.2f} kg")
    sheet.cell(row=summary_start_row, column=1).font = normal_font
    sheet.merge_cells(start_row=summary_start_row, start_column=1, end_row=summary_start_row, end_column=8)
    summary_start_row += 1

    # Export date
    from datetime import datetime
    sheet.cell(row=summary_start_row, column=1, value=f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    sheet.merge_cells(start_row=summary_start_row, start_column=1, end_row=summary_start_row, end_column=8)
    summary_start_row += 3  # Space before table

    # === TABLE STARTS HERE ===
    table_start_row = summary_start_row

    # Titles and field_attrs (unchanged)
    titles = [
        "SHIPMENT ID", "COMPANY", "VESSEL", "DOCS", "ORDER.NO", "SUPPLIER",
        "C.PACKING", "QTY", "UNIT", "SIZE", "WEIGHT", "IN DATE", "W/H",
        "BY", "IMP.BL", "EXP.BL", "PORT", "ONBOARD DATE",
        "REMARK", "DIVISION", "STATUS"
    ]
    field_attrs = [
        "shipment_id", "company", "vessel", "docs", "Order_No", "supplier",
        "c_packing", "quanty", "unit", "size", "weight", "in_date", "warehouse",
        "by", "IMP_BL", "EXP_BL", "port", "out_date",
        "remark", "division", "flag_status"
    ]

    # Column widths setup (unchanged)
    column_max_widths = {}
    MIN_WIDTH = 10
    CHAR_FACTOR = 1.3
    for i, title in enumerate(titles):
        column_max_widths[i + 1] = max(MIN_WIDTH, len(title) * CHAR_FACTOR)

    title_font = Font(bold=True, size=14)
    alignment = Alignment(horizontal='center', vertical='center')

    # Write Titles
    for i, title in enumerate(titles):
        col_idx = i + 1
        cell = sheet.cell(row=table_start_row, column=col_idx)
        cell.value = title
        cell.font = title_font
        cell.alignment = alignment

    # === DATA ROWS ===
    for j, shipment in enumerate(shipments_ticked):
        row_num = table_start_row + j + 1

        row_data = []
        for attr in field_attrs:
            value = getattr(shipment, attr, None)

            if isinstance(value, date):
                value_str = value.strftime('%Y-%m-%d')
            elif attr == "division":
                division_value = str(value) if value is not None else ""
                value_str = "D" if division_value.lower() in ["nan", ""] else division_value
            elif value is None or str(value).lower() in ['none', 'nan']:
                value_str = ""
            else:
                value_str = str(value)

            row_data.append(value_str)

        for col_idx, value_str in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_idx)
            cell.value = value_str
            cell.alignment = alignment

            content_width = len(value_str) * CHAR_FACTOR
            current_max = column_max_widths.get(col_idx, MIN_WIDTH)
            column_max_widths[col_idx] = max(current_max, content_width)

    # === APPLY COLUMN WIDTHS ===
    for col_idx, width in column_max_widths.items():
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = min(200, width + 2)

    wb.save(response)
    return response


def paginate_shipment_delete(request, context, ticked_nos=None):
    ticked_nos = ticked_nos or request.session.get("shipmentTickedNos", [])
    deleted_count = 0
    logger.info(f"paginate_shipment_delete: Request method: {request.method}, Ticked numbers: {ticked_nos}")
    if ticked_nos:
        try:
            with transaction.atomic():
                # Ensure ticked_nos are integers
                ticked_nos = [int(no) for no in ticked_nos if no]
                # Check if any shipments exist
                existing_count = Shipment.objects.filter(number__in=ticked_nos).count()
                if existing_count == 0:
                    logger.info(f"paginate_shipment_delete: No shipments found for ticked_nos: {ticked_nos}. Skipping deletion.")
                    return {"deleted_count": 0}

                # Perform deletion
                deleted_count = Shipment.objects.filter(number__in=ticked_nos).delete()[0]
                logger.info(f"paginate_shipment_delete: Deleted {deleted_count} shipments for ticked_nos: {ticked_nos}")
        except Exception as e:
            logger.error(f"paginate_shipment_delete: Deletion error: {str(e)}")
            raise
    else:
        logger.warning("paginate_shipment_delete: No ticked numbers provided")
    return {"deleted_count": deleted_count}

def apply_flag_status_filter(queryset, flag_values):
    """
    Manually apply flag_status filter with custom BLANK handling.
    flag_values: list of status values (e.g., ['START', 'COMPLETED', 'BLANK'])
    """
    if not flag_values:
        return queryset

    try:
        queryset = queryset.annotate(_trimmed_flag=Trim('flag_status'))
    except Exception:
        # If Trim unavailable, proceed without it
        pass

    q = Q()
    non_blank = [v for v in flag_values if v != 'BLANK']
    if 'BLANK' in flag_values:
        q |= Q(flag_status__exact='BLANK')
        q |= Q(_trimmed_flag='')
        q |= Q(flag_status__isnull=True)
    if non_blank:
        q |= Q(_trimmed_flag__in=non_blank)
    
    return queryset.filter(q)


def paginate_shipment_modify(
    request,
    context: dict,
    form_keyword: str="shipmentModForm",
    selections_keyword: str="allTickedShipments",
):
  shipments_ticked = context[selections_keyword]
  form_result = ShipmentModificationForm(request.POST)
  if form_result.is_valid():
    mod_in_date = form_result.cleaned_data.get('in_dateM')
    mod_out_date = form_result.cleaned_data.get('out_dateM')
    mod_job_number = form_result.cleaned_data.get('job_numberM')
    mod_port = form_result.cleaned_data.get('portM')
    mod_flag_status = form_result.cleaned_data.get('flag_statusM')
    mod_remark = form_result.cleaned_data.get('remarkM')
    mod_memo = form_result.cleaned_data.get('memoM')
    mod_warehouse = form_result.cleaned_data.get('warehouseM')
    mod_EXP_BL = form_result.cleaned_data.get('EXP_BLM')
    for shipment_modified in shipments_ticked:
      if mod_in_date:
        shipment_modified.in_date = mod_in_date
      if mod_out_date:
        shipment_modified.out_date = mod_out_date
      if mod_job_number:
        shipment_modified.job_number = mod_job_number
      if mod_port:
        shipment_modified.port = mod_port
      if mod_in_date:
        shipment_modified.in_date = mod_in_date
      if mod_flag_status:
        shipment_modified.flag_status = mod_flag_status
      if mod_remark:
        shipment_modified.remark = mod_remark
      if mod_memo:
        shipment_modified.memo = mod_memo
      if mod_warehouse:
        shipment_modified.warehouse = mod_warehouse
      if mod_EXP_BL:
        shipment_modified.EXP_BL = mod_EXP_BL

      shipment_modified.correct_org = request.user.userID
      shipment_modified.save()
    return context
